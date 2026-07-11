import openpyxl
from pathlib import Path

path = Path(r"C:\Users\pc\OneDrive\桌面\検品検针報告書SY26-000.xlsx")
wb = openpyxl.load_workbook(path, data_only=False)
print("SHEETS", wb.sheetnames)
for ws in wb.worksheets:
    print("SHEET", ws.title, ws.max_row, ws.max_column)
    print("MERGES", [str(rng) for rng in ws.merged_cells.ranges][:120])
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), min_col=1, max_col=min(ws.max_column, 30)):
        values = [cell.value for cell in row]
        if any(value is not None for value in values):
            print(row[0].row, values)
